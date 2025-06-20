from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import io
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv

from google.generativeai import GenerativeModel, configure
import json
import time
import re

# --- Cargar variables de entorno del archivo .env ---
load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

# --- Configurar la API de Gemini ---
configure(api_key=GEMINI_API_KEY)

app = Flask(__name__)
CORS(app)

# --- Definiciones de Excel y Clasificación ---
# Estas constantes deben ser idénticas a las de tu src/types.ts
DISPLAY_TO_INTERNAL_KEY_MAP = {
    "JURISDICCIÓN": "JURISDICCION",
    "CALIFICACIÓN": "CALIFICACION LEGAL",
    "MODALIDAD": "MODALIDAD",
    "VICTIMA/S": "VICTIMA",
    "IMPUTADOS": "IMPUTADO",
    "MENOR/MAYOR": "MAYOR O MENOR",
    "ARMAS": "ARMA",
    "LUGAR": "LUGAR",
    "TENTATIVA": "TENTATIVA",
    "OBSERVACIÓN": "OBSERVACION",
    "FRECUENCIA": "FRECUENCIA"
}

VALORES_VALIDOS = {
    "CALIFICACION LEGAL": ["ROBO", "HURTO", "LESIONES", "HOMICIDIO", "USURPACION", "ABUSO SEXUAL", "LEY 23737", "ABIGEATO", "ESTAFAS", "ABUSO DE ARMAS", "TENENCIA DE ARMAS", "PORTACION DE ARMAS", "ENCUBRIMIENTO", "NINGUNO DE INTERÉS", "OTROS"],
    "MODALIDAD": ["ASALTO", "MOTOCHORRO", "ENTRADERA", "VIOLENCIA DE GÉNERO", "HOMICIDIO SIMPLE", "FEMICIDIO", "INTRAFAMILIAR", "EN RIÑA", "EN OCASIÓN DE ROBO", "AJUSTE DE CUENTAS", "ENFRENTAMIENTO ARMADO", "SUSTRACCION AUTOMOTOR", "SUSTRACCION MOTOVEHICULO", "ABUSO SEXUAL SIMPLE", "ABUSO SEXUAL CON ACCESO CARNAL", "TENENCIA", "CONSUMO", "COMERCIALIZACION", "SIEMBRA", "ABIGEATO", "ESTAFA MARKETPLACE", "ESTAFA WHATSAPP", "ESTAFA CUENTO DEL TIO", "ESTAFA OTROS", "ABUSO DE ARMAS", "TENENCIA DE ARMAS", "PORTACION DE ARMAS", "ENCUBRIMIENTO", "NO ESPECIFICADO"],
    "ARMA": ["FUEGO", "BLANCA", "IMPROPIA", "NO ESPECIFICADO"],
    "LESIONADA": ["SI", "NO"],
    "VICTIMA": ["MASCULINO", "FEMENINO", "AMBOS", "NO ESPECIFICADO"],
    "IMPUTADO": ["MASCULINO", "FEMENINO", "AMBOS", "NO ESPECIFICADO"],
    "MAYOR O MENOR": ["MAYOR", "MENOR", "AMBOS", "NO ESPECIFICADO"],
    "JURISDICCION": ["JOSÉ C. PAZ", "SAN MIGUEL", "MALVINAS ARGENTINAS", "PILAR", "TRES DE FEBRERO", "MORENO", "RODRIGUEZ", "GENERAL PAZ", "NAVARRO", "MERCEDES", "SUIPACHA", "LUJAN", "GENERAL LAS HERAS", "MARCOS PAZ", "GENERAL RODRÍGUEZ", "EXALTACIÓN DE LA CRUZ", "CAMPANA", "ZÁRATE", "ESCOBAR", "TIGRE", "SAN FERNANDO", "VICENTE LÓPEZ", "SAN ISIDRO", "SAN MARTIN", "HURLINGHAM", "ITUZAINGÓ", "MERLO", "MORÓN", "LA MATANZA", "EZEIZA", "ESTEBAN ECHEVERRÍA", "LANÚS", "LOMAS DE ZAMORA", "AVELLANEDA", "QUILMES", "BERAZATEGUI", "FLORENCIO VARELA", "LA PLATA", "ENSENADA", "BERISSO", "BRANDSEN", "PRESIDENTE PERÓN", "SAN VICENTE", "CAÑUELAS", "GENERAL ALVEAR", "OLAVARRÍA", "AZUL", "TANDIL", "GENERAL PUEYRREDÓN", "MIRAMAR", "NECOCHEA", "BALCARCE", "OTRO", "NO ESPECIFICADO"],
    "LUGAR": ["FINCA", "VÍA PÚBLICA", "COMERCIO", "ESTABLECIMIENTO EDUCATIVO", "TRANSPORTE PÚBLICO", "BANCO", "HOSPITAL", "OTRO", "NO ESPECIFICADO"],
    "TENTATIVA": ["SI", "NO"],
    "OBSERVACION": ["NO ESPECIFICADO"],
    "FRECUENCIA": ["DIARIA", "SEMANAL", "MENSUAL", "OCASIONAL", "NO ESPECIFICADO"]
}

CLASIFICACION_DEFAULT_BACKEND = {
    "id_hecho": "", "nro_registro": "", "ipp": "", "fecha_carga": "", "hora_carga": "",
    "dependencia": "", "fecha_inicio_hecho": "", "hora_inicio_hecho": "",
    "partido_hecho": "", "localidad_hecho": "", "latitud": "", "calle": "",
    "longitud": "", "altura": "", "entre": "", "calificaciones": "",
    "relato": "",
    "JURISDICCIÓN": "NO ESPECIFICADO", "CALIFICACIÓN": "NINGUNO DE INTERÉS",
    "MODALIDAD": "NO ESPECIFICADO", "ARMA": "NO ESPECIFICADO", "LESIONADA": "NO",
    "VICTIMA/S": "NO ESPECIFICADO", "NO": "", "IMPUTADOS": "NO ESPECIFICADO",
    "MENOR/MAYOR": "NO ESPECIFICADO", "LUGAR": "NO ESPECIFICADO",
    "TENTATIVA": "NO", "OBSERVACIÓN": "NO ESPECIFICADO", "FRECUENCIA": "NO ESPECIFICADO"
}

EXCEL_OUTPUT_HEADERS = [
    "id_hecho", "nro_registro", "ipp", "fecha_carga", "hora_carga", "dependencia",
    "fecha_inicio_hecho", "hora_inicio_hecho", "partido_hecho", "localidad_hecho",
    "latitud", "calle", "longitud", "altura", "entre", "calificaciones",
    "relato", "JURISDICCIÓN", "CALIFICACIÓN", "MODALIDAD", "VICTIMA/S", "NO",
    "IMPUTADOS", "MENOR/MAYOR", "ARMAS", "LUGAR", "TENTATIVA", "OBSERVACIÓN", "FRECUENCIA"
]

IA_CLASSIFICATION_KEYS_DISPLAY = [
    "JURISDICCIÓN", "CALIFICACIÓN", "MODALIDAD", "VICTIMA/S", "IMPUTADOS",
    "MENOR/MAYOR", "ARMAS", "LUGAR", "TENTATIVA", "OBSERVACIÓN", "FRECUENCIA"
]

def validate_and_clean_classification(raw_classification, log_entries):
    clean_classification = dict(CLASIFICACION_DEFAULT_BACKEND)

    for display_key in IA_CLASSIFICATION_KEYS_DISPLAY:
        internal_key = DISPLAY_TO_INTERNAL_KEY_MAP.get(display_key, display_key)
        raw_value = raw_classification.get(internal_key)

        valid_options = VALORES_VALIDOS.get(internal_key, [])

        if raw_value is not None and isinstance(raw_value, str):
            upper_case_value = raw_value.upper().strip()

            if valid_options and upper_case_value in valid_options:
                clean_classification[display_key] = upper_case_value
            else:
                clean_classification[display_key] = CLASIFICACION_DEFAULT_BACKEND.get(display_key, "NO ESPECIFICADO")
                log_entries.append(f"    [!] ADVERTENCIA: Campo '{display_key}'. Valor de IA '{raw_value}' NO VÁLIDO. Usando '{clean_classification[display_key]}'.")
        else:
            clean_classification[display_key] = CLASIFICACION_DEFAULT_BACKEND.get(display_key, "NO ESPECIFICADO")
            if raw_value is not None:
                log_entries.append(f"    [!] ADVERTENCIA: Campo '{display_key}'. Valor de IA '{raw_value}' (tipo {type(raw_value).__name__}) NO ES VÁLIDO. Usando '{clean_classification[display_key]}'.")
    return clean_classification

@app.route('/process-excel', methods=['POST'])
def process_excel_backend():
    log_entries = []
    try:
        if 'dataFile' not in request.files or 'templateFile' not in request.files:
            log_entries.append("ERROR: Ambos archivos (datos y plantilla) son requeridos.")
            return jsonify({"error": "Ambos archivos (datos y plantilla) son requeridos."}), 400

        data_file = request.files['dataFile']
        template_file = request.files['templateFile']

        # 1. Leer el archivo de datos
        data_wb = openpyxl.load_workbook(io.BytesIO(data_file.read()))
        data_ws = data_wb.active
        
        # Iniciar la lista de filas leídas del archivo de datos
        data_rows = [] 
        data_headers = [cell.value for cell in data_ws[1]]
        for r_idx in range(2, data_ws.max_row + 1):
            row_data = {}
            for c_idx, header in enumerate(data_headers):
                cell_value = data_ws.cell(row=r_idx, column=c_idx + 1).value
                row_data[header] = str(cell_value).strip() if cell_value is not None else ""
            data_rows.append(row_data)

        if not data_rows:
            log_entries.append("ERROR: El archivo de datos está vacío o no tiene filas válidas.")
            return jsonify({"error": "El archivo de datos está vacío o no tiene filas válidas."}), 400

        # 2. Leer la plantilla de Excel
        template_wb = openpyxl.load_workbook(io.BytesIO(template_file.read()))
        template_ws = template_wb.active

        # 3. Identificar la fila de inicio de datos en la plantilla
        start_row_in_template = 3

        model = GenerativeModel("gemini-1.5-flash-latest")

        # Inicia `current_template_row` ANTES del bucle
        current_template_row = start_row_in_template # <--- SOLUCIÓN (Asegura que esté definida)

        for i, row_data_from_file in enumerate(data_rows): # ¡CORREGIDO AQUÍ: Usar 'data_rows'!
            # Calcula `current_template_row` en cada iteración del bucle
            current_template_row = start_row_in_template + i # Esta asignación ocurre en cada iteración

            log_entries.append(f"\n--- PROCESANDO FILA {i + 2} (ID Original: {row_data_from_file.get('id_hecho', 'N/A')}) ---\n")
            
            relato_text = row_data_from_file.get('relato') or row_data_from_file.get('RELATO') or row_data_from_file.get('RELATO ORIGINAL') or ""
            relato_text = relato_text.strip()

            log_entries.append(f"  Relato Original (Extracto): '{relato_text[:150]}{'...' if len(relato_text) > 150 else ''}'\n")

            if not relato_text:
                log_entries.append(f"  >>> RESULTADO: RELATO VACÍO - Clasificación por defecto aplicada. <<<\n")
                for col_idx, header in enumerate(EXCEL_OUTPUT_HEADERS):
                    value_to_write = CLASIFICACION_DEFAULT_BACKEND.get(header, "")
                    template_ws.cell(row=current_template_row, column=col_idx + 1).value = value_to_write
                continue

            prompt = f"""Analiza el siguiente relato delictivo. Identifica la información relevante para las siguientes categorías.
            Tu respuesta debe ser estricta y ÚNICAMENTE un objeto JSON. No incluyas ningún otro texto, markdown, preámbulos o explicaciones fuera del JSON.

            Relato: "{relato_text}"

            INSTRUCCIONES CLAVE DE CLASIFICACIÓN:
            - Para cada categoría, DEBES seleccionar uno de los valores válidos de la lista proporcionada. Si un valor no se puede determinar con certeza a partir del relato, o el relato no ofrece información para una categoría, DEBES utilizar el valor por defecto "NO ESPECIFICADO" (o "NO" para campos binarios como LESIONADA/TENTATIVA) de la lista de VALORES VÁLIDOS.
            - Los valores deben ser EXACTOS (MAYÚSCULAS y SIN ACENTOS si el valor en la lista no los tiene, o CON ACENTOS si la lista los tiene).
            - Para "CALIFICACION", prioriza el delito principal. Si no es de interés, usa "NINGUNO DE INTERES".
            - Para "ARMAS", si no es de fuego o blanca, y se usó un objeto para dañar, clasifica como "IMPROPIA". Si no se menciona arma o no se usó, usa "NO ESPECIFICADO".
            - Para "LESIONADA", responde "SI" si el relato menciona lesiones/heridas, "NO" en caso contrario.
            - Para "VICTIMA", "IMPUTADO", si hay varios géneros o no se especifica, usa "AMBOS" o "NO ESPECIFICADO" respectivamente.
            - Para "MENOR/MAYOR", busca indicios de edad del imputado. Si no se puede determinar, usa "NO ESPECIFICADO".
            - Para "JURISDICCIÓN" y "LUGAR", intenta extraer del relato los valores más precisos de las listas dadas.
            - Para "TENTATIVA", "SI" si el delito fue intentado pero no consumado, "NO" si fue consumado o no aplica.
            - Para "OBSERVACION" y "FRECUENCIA", usa "NO ESPECIFICADO" a menos que el relato brinde información explícita para una de las opciones válidas.

            LISTA DE VALORES VÁLIDOS POR CAMPO (elige uno EXACTO, en MAYÚSCULAS):
            "CALIFICACION LEGAL": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["CALIFICACION LEGAL"]])}]
            "MODALIDAD": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["MODALIDAD"]])}]
            "ARMA": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["ARMA"]])}]
            "LESIONADA": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["LESIONADA"]])}]
            "VICTIMA": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["VICTIMA"]])}]
            "IMPUTADO": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["IMPUTADO"]])}]
            "MAYOR O MENOR": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["MAYOR O MENOR"]])}]
            "JURISDICCION": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["JURISDICCION"]])}]
            "LUGAR": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["LUGAR"]])}]
            "TENTATIVA": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["TENTATIVA"]])}]
            "OBSERVACION": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["OBSERVACION"]])}]
            "FRECUENCIA": [{", ".join([f'"{v}"' for v in VALORES_VALIDOS["FRECUENCIA"]])}]
            """

            try:
                response = model.generate_content(prompt)
                raw_text_response = response.text
                escaped_raw_text = raw_text_response.replace("'", "\\'").replace("\n", "\\n")
                log_entries.append(f"  Respuesta Cruda de IA (para depuración):\n  '{escaped_raw_text}'\n")

                json_match = None
                match = re.search(r'\{[\s\S]*\}', raw_text_response)
                if match:
                    json_match = match.group(0)

                classified_by_ia = {}
                if json_match:
                    try:
                        classified_by_ia = json.loads(json_match)
                    except json.JSONDecodeError as e:
                        log_entries.append(f"    [!!!] ERROR: Falló el parsing JSON de la IA: {e}. Respuesta IA: {raw_text_response}\n")
                        classified_by_ia = {}

                validated_classification = validate_and_clean_classification(classified_by_ia, log_entries)
                log_entries.append(f"  >>> RESULTADO: Procesado con IA. <<<\n")
                
                # Escribir en la plantilla de Excel
                for col_idx, header in enumerate(EXCEL_OUTPUT_HEADERS):
                    cell_value_to_write = ""
                    if header in validated_classification:
                        cell_value_to_write = validated_classification[header]
                    elif header in row_data_from_file:
                        cell_value_to_write = row_data_from_file[header]
                    elif header == "relato":
                        cell_value_to_write = relato_text
                    else:
                        cell_value_to_write = CLASIFICACION_DEFAULT_BACKEND.get(header, "")

                    template_ws.cell(row=current_template_row, column=col_idx + 1).value = cell_value_to_write

                log_entries.append(f"  Clasificación Final (IA + Originales):\n{json.dumps(validated_classification, indent=2)}\n")

            except Exception as e:
                user_friendly_error_message = f"Error de IA al procesar: {str(e)}"
                if "429 You exceeded your current quota" in str(e):
                    user_friendly_error_message = "Error de Cuota de la API de Gemini: Límite de uso excedido para esta clave. Por favor, espere 24 horas o contacte al administrador para más cuota."
                elif "500" in str(e) or "503" in str(e):
                    user_friendly_error_message = "Error interno del servicio de IA de Gemini. Por favor, intente de nuevo más tarde."
                
                log_entries.append(f"    [!!!] ERROR GENERAL EN IA: {user_friendly_error_message}. No se pudo clasificar esta fila.\n")
                # Llenar con valores por defecto y copiar originales
                for col_idx, header in enumerate(EXCEL_OUTPUT_HEADERS):
                    value_to_write = CLASIFICACION_DEFAULT_BACKEND.get(header, "")
                    if header == "relato":
                        value_to_write = relato_text
                    elif header in row_data_from_file:
                        value_to_write = row_data_from_file[header]
                    template_ws.cell(row=current_template_row, column=col_idx + 1).value = value_to_write
                log_entries.append(f"  Clasificación Final (Usando solo Originales y Defaults):\n{json.dumps(CLASIFICACION_DEFAULT_BACKEND, indent=2)}\n")

            log_entries.append(f"-----------------------------------------------------------------\n")

            if (i + 1) % 50 == 0 and i < len(data_rows) - 1: # ¡CORREGIDO AQUÍ: Usar 'data_rows'!
                log_entries.append(f"\n[ PAUSA: Esperando 0.5 segundos para respetar la cuota de la API. (Después de {i + 1} filas) ]\n\n")
                time.sleep(0.5)

        log_entries.append(f"\n=================================================================\n")
        log_entries.append(f"= PROCESAMIENTO EN BACKEND FINALIZADO ({len(data_rows)} filas) =\n") # ¡CORREGIDO AQUÍ: Usar 'data_rows'!
        log_entries.append(f"=================================================================\n")


        # Guardar la plantilla modificada en memoria y enviarla de vuelta
        output_buffer = io.BytesIO()
        template_wb.save(output_buffer)
        output_buffer.seek(0)

        return send_file(
            output_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Clasificado_Final_IA.xlsx'
        )

    except Exception as e:
        error_message = f"Error en el procesamiento del servidor: {str(e)}"
        log_entries.append(f"\n=================================================================\n")
        log_entries.append(f"= ERROR CRÍTICO EN EL BACKEND =\n")
        log_entries.append(f"=================================================================\n")
        log_entries.append(f"Mensaje: {error_message}\n")
        log_entries.append(f"Fecha y Hora: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())}\n")
        
        print(error_message)

        return jsonify({"error": error_message, "log_details": "\n".join(log_entries)}), 500


# Punto de entrada para el servidor Flask
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)